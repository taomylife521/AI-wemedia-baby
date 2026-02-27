"""
导出工具模块
文件路径：src/utils/export_utils.py
功能：提供任务配置和结果导出功能
"""

import json
import csv
from typing import List, Dict, Any, Optional
from pathlib import Path
import logging

logger = logging.getLogger(__name__)


def export_task_config(task: Dict[str, Any], file_path: str) -> bool:
    """导出任务配置为JSON文件
    
    Args:
        task: 任务信息字典
        file_path: 导出文件路径
    
    Returns:
        如果导出成功返回True，否则返回False
    """
    try:
        # 准备导出数据
        export_data = {
            'task_name': task.get('task_name'),
            'task_description': task.get('task_description'),
            'platform_username': task.get('platform_username'),
            'platform': task.get('platform'),
            'task_type': task.get('task_type'),
            'script_config': task.get('script_config', {}),
            'video_count': task.get('video_count', 0),
            'retry_count': task.get('retry_count', 3),
            'delay_seconds': task.get('delay_seconds', 5),
            'max_concurrent': task.get('max_concurrent', 1),
            'created_at': task.get('created_at')
        }
        
        # 写入JSON文件
        with open(file_path, 'w', encoding='utf-8') as f:
            json.dump(export_data, f, ensure_ascii=False, indent=2)
        
        logger.info(f"导出任务配置成功: {file_path}")
        return True
    except Exception as e:
        logger.error(f"导出任务配置失败: {e}", exc_info=True)
        return False


def export_task_results(
    task: Dict[str, Any],
    executions: List[Dict[str, Any]],
    file_path: str,
    format: str = 'csv'
) -> bool:
    """导出任务结果为CSV或Excel文件
    
    Args:
        task: 任务信息字典
        executions: 执行记录列表
        file_path: 导出文件路径
        format: 导出格式（csv/excel）
    
    Returns:
        如果导出成功返回True，否则返回False
    """
    try:
        if format == 'csv':
            return _export_to_csv(task, executions, file_path)
        elif format == 'excel':
            return _export_to_excel(task, executions, file_path)
        else:
            logger.error(f"不支持的导出格式: {format}")
            return False
    except Exception as e:
        logger.error(f"导出任务结果失败: {e}", exc_info=True)
        return False


def _export_to_csv(
    task: Dict[str, Any],
    executions: List[Dict[str, Any]],
    file_path: str
) -> bool:
    """导出为CSV格式"""
    try:
        with open(file_path, 'w', encoding='utf-8-sig', newline='') as f:
            writer = csv.writer(f)
            
            # 写入任务信息
            writer.writerow(['任务信息'])
            writer.writerow(['任务名称', task.get('task_name', '')])
            writer.writerow(['账号', task.get('platform_username', '')])
            writer.writerow(['平台', task.get('platform', '')])
            writer.writerow(['总数量', task.get('video_count', 0)])
            writer.writerow(['已完成', task.get('completed_count', 0)])
            writer.writerow(['失败数', task.get('failed_count', 0)])
            writer.writerow(['创建时间', task.get('created_at', '')])
            writer.writerow([])
            
            # 写入执行记录表头
            writer.writerow(['执行记录'])
            writer.writerow([
                '序号', '文件路径', '标题', '状态', '错误信息',
                '重试次数', '发布URL', '开始时间', '完成时间'
            ])
            
            # 写入执行记录数据
            for execution in executions:
                writer.writerow([
                    execution.get('execution_index', ''),
                    execution.get('file_path', ''),
                    execution.get('title', ''),
                    execution.get('status', ''),
                    execution.get('error_message', ''),
                    execution.get('retry_count', 0),
                    execution.get('publish_url', ''),
                    execution.get('started_at', ''),
                    execution.get('completed_at', '')
                ])
        
        logger.info(f"导出CSV成功: {file_path}")
        return True
    except Exception as e:
        logger.error(f"导出CSV失败: {e}", exc_info=True)
        return False


def _export_to_excel(
    task: Dict[str, Any],
    executions: List[Dict[str, Any]],
    file_path: str
) -> bool:
    """导出为Excel格式（需要openpyxl库）"""
    try:
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, Alignment
        except ImportError:
            logger.error("导出Excel需要openpyxl库，请先安装: pip install openpyxl")
            return False
        
        wb = Workbook()
        ws = wb.active
        ws.title = "任务结果"
        
        # 写入任务信息
        row = 1
        ws.cell(row, 1, "任务信息").font = Font(bold=True, size=14)
        row += 1
        
        info_labels = [
            ('任务名称', task.get('task_name', '')),
            ('账号', task.get('platform_username', '')),
            ('平台', task.get('platform', '')),
            ('总数量', task.get('video_count', 0)),
            ('已完成', task.get('completed_count', 0)),
            ('失败数', task.get('failed_count', 0)),
            ('创建时间', task.get('created_at', ''))
        ]
        
        for label, value in info_labels:
            ws.cell(row, 1, label).font = Font(bold=True)
            ws.cell(row, 2, value)
            row += 1
        
        row += 1
        
        # 写入执行记录表头
        headers = [
            '序号', '文件路径', '标题', '状态', '错误信息',
            '重试次数', '发布URL', '开始时间', '完成时间'
        ]
        
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row, col, header)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center')
        
        row += 1
        
        # 写入执行记录数据
        for execution in executions:
            ws.cell(row, 1, execution.get('execution_index', ''))
            ws.cell(row, 2, execution.get('file_path', ''))
            ws.cell(row, 3, execution.get('title', ''))
            ws.cell(row, 4, execution.get('status', ''))
            ws.cell(row, 5, execution.get('error_message', ''))
            ws.cell(row, 6, execution.get('retry_count', 0))
            ws.cell(row, 7, execution.get('publish_url', ''))
            ws.cell(row, 8, execution.get('started_at', ''))
            ws.cell(row, 9, execution.get('completed_at', ''))
            row += 1
        
        # 调整列宽
        ws.column_dimensions['A'].width = 10
        ws.column_dimensions['B'].width = 50
        ws.column_dimensions['C'].width = 30
        ws.column_dimensions['D'].width = 15
        ws.column_dimensions['E'].width = 50
        ws.column_dimensions['F'].width = 12
        ws.column_dimensions['G'].width = 50
        ws.column_dimensions['H'].width = 20
        ws.column_dimensions['I'].width = 20
        
        wb.save(file_path)
        logger.info(f"导出Excel成功: {file_path}")
        return True
    except Exception as e:
        logger.error(f"导出Excel失败: {e}", exc_info=True)
        return False


def export_multiple_tasks(
    tasks: List[Dict[str, Any]],
    file_path: str,
    format: str = 'json'
) -> bool:
    """批量导出多个任务配置
    
    Args:
        tasks: 任务列表
        file_path: 导出文件路径
        format: 导出格式（json/csv）
    
    Returns:
        如果导出成功返回True，否则返回False
    """
    try:
        if format == 'json':
            with open(file_path, 'w', encoding='utf-8') as f:
                json.dump(tasks, f, ensure_ascii=False, indent=2)
        elif format == 'csv':
            with open(file_path, 'w', encoding='utf-8-sig', newline='') as f:
                writer = csv.writer(f)
                writer.writerow([
                    '任务ID', '任务名称', '账号', '平台', '类型',
                    '总数量', '已完成', '失败数', '状态', '创建时间'
                ])
                for task in tasks:
                    writer.writerow([
                        task.get('id', ''),
                        task.get('task_name', ''),
                        task.get('platform_username', ''),
                        task.get('platform', ''),
                        task.get('task_type', ''),
                        task.get('video_count', 0),
                        task.get('completed_count', 0),
                        task.get('failed_count', 0),
                        task.get('status', ''),
                        task.get('created_at', '')
                    ])
        else:
            logger.error(f"不支持的导出格式: {format}")
            return False
        
        logger.info(f"批量导出任务成功: {file_path}, 数量={len(tasks)}")
        return True
    except Exception as e:
        logger.error(f"批量导出任务失败: {e}", exc_info=True)
        return False


def export_publish_records(
    records: List[Dict[str, Any]],
    file_path: str,
    format: str = 'csv'
) -> bool:
    """导出发布记录为CSV或Excel文件
    
    Args:
        records: 发布记录列表
        file_path: 导出文件路径
        format: 导出格式（csv/excel）
    
    Returns:
        如果导出成功返回True，否则返回False
    """
    try:
        if format == 'csv':
            return _export_publish_records_to_csv(records, file_path)
        elif format == 'excel':
            return _export_publish_records_to_excel(records, file_path)
        else:
            logger.error(f"不支持的导出格式: {format}")
            return False
    except Exception as e:
        logger.error(f"导出发布记录失败: {e}", exc_info=True)
        return False


def _export_publish_records_to_csv(
    records: List[Dict[str, Any]],
    file_path: str
) -> bool:
    """导出发布记录为CSV格式"""
    try:
        import os
        
        with open(file_path, 'w', encoding='utf-8-sig', newline='') as f:
            writer = csv.writer(f)
            
            # 写入表头
            writer.writerow([
                'ID', '平台', '账号', '文件路径', '文件类型', '标题', '描述',
                '标签', '状态', '错误信息', '发布链接', '创建时间', '更新时间'
            ])
            
            # 写入数据
            for record in records:
                # 格式化标签
                tags = record.get('tags', '')
                if tags:
                    try:
                        import json
                        if isinstance(tags, str):
                            tags_list = json.loads(tags)
                        else:
                            tags_list = tags
                        tags_str = ', '.join(tags_list) if isinstance(tags_list, list) else str(tags)
                    except:
                        tags_str = str(tags)
                else:
                    tags_str = ''
                
                # 平台显示名称
                platform = record.get('platform', '')
                platform_display = {
                    'douyin': '抖音',
                    'kuaishou': '快手',
                    'xiaohongshu': '小红书'
                }.get(platform, platform)
                
                # 状态显示名称
                status = record.get('status', '')
                status_display = {
                    'success': '成功',
                    'failed': '失败',
                    'pending': '待发布'
                }.get(status, status)
                
                writer.writerow([
                    record.get('id', ''),
                    platform_display,
                    record.get('platform_username', ''),
                    record.get('file_path', ''),
                    record.get('file_type', ''),
                    record.get('title', ''),
                    record.get('description', ''),
                    tags_str,
                    status_display,
                    record.get('error_message', ''),
                    record.get('publish_url', ''),
                    record.get('created_at', ''),
                    record.get('updated_at', '') or ''
                ])
        
        logger.info(f"导出发布记录CSV成功: {file_path}, 数量={len(records)}")
        return True
    except Exception as e:
        logger.error(f"导出发布记录CSV失败: {e}", exc_info=True)
        return False


def _export_publish_records_to_excel(
    records: List[Dict[str, Any]],
    file_path: str
) -> bool:
    """导出发布记录为Excel格式"""
    try:
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, Alignment
        except ImportError:
            logger.error("导出Excel需要openpyxl库，请先安装: pip install openpyxl")
            return False
        
        wb = Workbook()
        ws = wb.active
        ws.title = "发布记录"
        
        # 写入表头
        headers = [
            'ID', '平台', '账号', '文件路径', '文件类型', '标题', '描述',
            '标签', '状态', '错误信息', '发布链接', '创建时间', '更新时间'
        ]
        
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(1, col, header)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center')
        
        # 写入数据
        for row, record in enumerate(records, start=2):
            # 格式化标签
            tags = record.get('tags', '')
            if tags:
                try:
                    import json
                    if isinstance(tags, str):
                        tags_list = json.loads(tags)
                    else:
                        tags_list = tags
                    tags_str = ', '.join(tags_list) if isinstance(tags_list, list) else str(tags)
                except:
                    tags_str = str(tags)
            else:
                tags_str = ''
            
            # 平台显示名称
            platform = record.get('platform', '')
            platform_display = {
                'douyin': '抖音',
                'kuaishou': '快手',
                'xiaohongshu': '小红书'
            }.get(platform, platform)
            
            # 状态显示名称
            status = record.get('status', '')
            status_display = {
                'success': '成功',
                'failed': '失败',
                'pending': '待发布'
            }.get(status, status)
            
            ws.cell(row, 1, record.get('id', ''))
            ws.cell(row, 2, platform_display)
            ws.cell(row, 3, record.get('platform_username', ''))
            ws.cell(row, 4, record.get('file_path', ''))
            ws.cell(row, 5, record.get('file_type', ''))
            ws.cell(row, 6, record.get('title', ''))
            ws.cell(row, 7, record.get('description', ''))
            ws.cell(row, 8, tags_str)
            ws.cell(row, 9, status_display)
            ws.cell(row, 10, record.get('error_message', ''))
            ws.cell(row, 11, record.get('publish_url', ''))
            ws.cell(row, 12, record.get('created_at', ''))
            ws.cell(row, 13, record.get('updated_at', '') or '')
        
        # 调整列宽
        column_widths = [8, 10, 15, 40, 12, 30, 40, 30, 10, 40, 40, 20, 20]
        for col, width in enumerate(column_widths, start=1):
            ws.column_dimensions[chr(64 + col)].width = width
        
        wb.save(file_path)
        logger.info(f"导出发布记录Excel成功: {file_path}, 数量={len(records)}")
        return True
    except Exception as e:
        logger.error(f"导出发布记录Excel失败: {e}", exc_info=True)
        return False
